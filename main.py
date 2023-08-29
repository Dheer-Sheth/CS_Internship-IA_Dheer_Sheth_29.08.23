from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import Tk, StringVar,OptionMenu, Button, Label, Entry, ttk
import barcode
from tkcalendar import Calendar, DateEntry
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont
import settings

from tkinter import messagebox
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Image




def read_barcodes_from_excel(file_path, sheet_name):
    workbook= load_workbook(file_path)
    sheet=workbook[sheet_name]
    barcodes=["                       "]
    for row in sheet.iter_rows(values_only=True):
        barcode_value=row[0]
        barcodes.append(barcode_value)
    return barcodes

def generate_barcode(selected_barcode, label_size):
    Code128= barcode.get_barcode_class('code128')
    code128= Code128(selected_barcode, writer=ImageWriter())
    barcode_image= code128.render()
    return barcode_image

def generate_pdf_with_barcodes(barcodes,label_size, labels_per_page):
    if label_size == "100mm x 150mm":
        label_width, label_height = 100, 150
    elif label_size == "75mm x 75mm":
        label_width, label_height = 75, 75

    doc = SimpleDocTemplate("barcode_labels.pdf", pagesize=(label_width, label_height), leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0)
    story = []

    for barcode_value in barcodes:
        barcode_image_path = generate_barcode(barcode_value, label_size)
        image = Image(barcode_image_path, width=label_width, height=label_height)
        story.append(image)

        if len(story) == labels_per_page:
            doc.build(story)
            story = []

    if story:
        doc.build(story)

    print("PDF generated with barcode labels")



def generate_selected_barcode():
    selected_barcode = selected_barcode_var.get()
    selected_date = date_entry.get_date().strftime("%Y-%m-%d")
    quantity = int(quantity_entry.get())
    vendor = vendor_entry.get()
    supplier = supplier_entry.get()

    for i in range(quantity):
        barcode_image = generate_barcode(selected_barcode, label_size)

        barcode_image = barcode_image.convert("RGB")


        barcode_with_date = Image.new('RGB', (barcode_image.width, barcode_image.height + 30), 'white')
        barcode_with_date.paste(barcode_image, (0, 0))

        draw = ImageDraw.Draw(barcode_with_date)
        date_text = f"Date: {selected_date}"
        text_width, text_height = draw.textsize(date_text)
        text_x = (barcode_image.width - text_width) // 2
        text_y = barcode_image.height + 5
        draw.text((text_x, text_y), date_text,  fill='black')

        save_filename = f'barcode_{selected_barcode}_{i + 1}.png'
        barcode_with_date.save(save_filename)

    selected_barcodes = [selected_barcode_var.get()]
    labels_per_page = int(settings.selected_labels_per_page.var.get())
    label_size = settings.selected_label_size_var.get()

    generate_pdf_with_barcodes(selected_barcodes, label_size, labels_per_page)




def add_item_callback():
    selected_barcode = selected_barcode_var.get()
    selected_date= date_entry.get_date().strftime("%Y-%m-%d")
    quantity = int(quantity_entry.get())
    vendor = vendor_entry.get()
    supplier = supplier_entry.get()

    for item in treeview.get_children():
        barcode_value = treeview.item(item)['values'][0]
        if barcode_value == selected_barcode:
            return

    treeview.insert("", "end", values=(selected_barcode, selected_date, quantity, vendor, supplier))


def show_settings_window():
    notebook.select(settings_frame)

root=Tk()
root.title('Barcode Dropdown')
root.geometry('1200x1200')



notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True)

main_frame = ttk.Frame(notebook)
notebook.add(main_frame, text="Main Window")

settings_frame = ttk.Frame(notebook)
notebook.add(settings_frame, text="Settings")



barcodes = read_barcodes_from_excel('data_barcodes.xlsx', 'Sheet1')

selected_barcode_var = StringVar(root)
selected_barcode_var.set(barcodes[0])

##login_button = Button(root, text="Login", command=open_login_window)
##login_button.pack(padx=10, pady=10)

selected_barcode_label= Label(main_frame, text="Select Barcode", font=('Helvatical Bold',15))
selected_barcode_label.pack(anchor= 'w',  padx=10, pady=5)

barcode_dropdown = OptionMenu(main_frame, selected_barcode_var, *barcodes)
barcode_dropdown.config(width=20)
barcode_dropdown.pack(anchor= 'w',  padx=10, pady=5)

vendor_label= Label(main_frame, text="Vendor Name" ,font=('Helvatical Bold',15))
vendor_label.pack(anchor = 'w',  padx=10, pady=5)

vendor_entry = Entry(main_frame)
vendor_entry.config(width=20)
vendor_entry.pack(anchor= 'w', padx=10, pady=5)

supplier_label = Label(main_frame, text="Supplier Name", font=('Helvatical Bold',15))
supplier_label.pack(anchor = 'w',  padx=10, pady=5)

supplier_entry = Entry(main_frame)
supplier_entry.config(width=20)
supplier_entry.pack(anchor = 'w',  padx=10, pady=5)


quantity_label= Label(main_frame, text="Enter Quantity", font=('Helvatical Bold',15))
quantity_label.pack(anchor = 'w' ,  padx=10, pady=5)

quantity_entry = Entry(main_frame)
quantity_entry.config(width=20)
quantity_entry.pack(anchor = 'w',  padx=10, pady=5)

date_label = Label(main_frame, text="Select Date", font=('Helvatical Bold',15))
date_label.pack(anchor = 'w',   padx=10, pady=5)

date_entry = DateEntry(main_frame, date_pattern="yyyy-mm-dd")
date_entry.config(width=20)
date_entry.pack(anchor = 'w',  padx=10, pady=5)

current_label_size_label = Label(main_frame, text=f"Current Label Size: {settings.selected_label_size}")
current_label_size_label.pack(anchor='w', padx=20, pady=5)


generate_button= Button(main_frame, text="Generate Barcode", command= generate_selected_barcode)
generate_button.pack(anchor = 'w',   padx=10, pady=10)


add_item_button = Button(main_frame, text="Add Item", command=add_item_callback)
add_item_button.pack(anchor = 'w',  padx=10, pady=10)

treeview = ttk.Treeview(main_frame, columns=("Barcode", "Date", "Quantity", "Vendor", "Supplier"), show="headings")
treeview.heading("Barcode", text="Barcode")
treeview.heading("Date", text="Date")
treeview.heading("Quantity", text="Quantity")
treeview.heading("Vendor", text="Vendor")
treeview.heading("Supplier", text="Supplier")
treeview.pack(anchor = 'w',  padx=20, pady=10)



selected_label_size_var = StringVar(settings_frame)
selected_label_size_var.set(settings.selected_label_size)


label_size_dropdown = OptionMenu(settings_frame, selected_label_size_var, "100mm x 150mm", "75mm x 75mm")
label_size_dropdown.pack(padx=20, pady=5)


apply_label_size_button = Button(settings_frame, text="Apply Label Size", command=settings.apply_label_size)
apply_label_size_button.pack(padx=20, pady=5)

selected_labels_per_page_var = StringVar(settings_frame)
selected_labels_per_page_var.set(settings.selected_labels_per_page)

labels_per_page_dropdown = OptionMenu(settings_frame, selected_labels_per_page_var, "1", "2", "4")
labels_per_page_dropdown.pack(padx=20, pady=5)


apply_labels_per_page_button = Button(settings_frame, text="Apply Labels per Page", command=settings.apply_labels_per_page)
apply_labels_per_page_button.pack(padx=20, pady=5)






root.mainloop()
